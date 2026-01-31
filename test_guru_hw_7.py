import zipfile
import os
import pytest
from pypdf import PdfReader
from openpyxl import load_workbook

resource_path = os.path.join(os.getcwd(), "resource")
pdf_path = os.path.join(resource_path , "Python Testing with Pytest (Brian Okken).pdf") # склеиваем путь к файлу readme2.rst
xlsx_path = os.path.join(resource_path , "file_example_XLSX_50.xlsx")
csv_path = os.path.join(resource_path , "data.csv")
zip_path = os.path.join(resource_path , "zipfile.zip")


@pytest.fixture
def test_zip_archive():
    with zipfile.ZipFile(zip_path, 'w') as zf:
        for file in (pdf_path, xlsx_path, csv_path):
            zf.write(file, os.path.basename(file))
    yield
    os.remove(zip_path)


def test_check_pdf_file(test_zip_archive):
    with zipfile.ZipFile(zip_path, "r") as zp:
        with zp.open("Python Testing with Pytest (Brian Okken).pdf") as file:
            reader = PdfReader(file)

            # проверка 1
            number_of_pages = len(reader.pages)  # узнаем количество страниц в файле
            assert number_of_pages == 256, 'Ошибка в количестве страниц в файле Pdf'

            # проверка 2
            page = reader.pages[1]
            text = page.extract_text()
            assert """Version: P1.0 (September 2017)""" in text, 'Ошибка в тексте файла Pdf'


def test_check_csv_file(test_zip_archive):
    with zipfile.ZipFile(zip_path, "r") as zp:
        with zp.open("data.csv") as file:
            text = file.readline().decode('utf-8')
            assert "policyID" in text, 'Ошибка в тексте файла csv'


def test_check_xlsx_file(test_zip_archive):
    with zipfile.ZipFile(zip_path, "r") as zp:
        with zp.open("file_example_XLSX_50.xlsx") as file:
            text = load_workbook(file)
            sheet = text.active
            assert sheet.cell(row=3, column=8).value == 1582